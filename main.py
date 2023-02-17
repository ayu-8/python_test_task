#!/usr/bin/python3
# Python 3.11.0

import traceback
import requests
import datetime
import json
import smtplib
import ssl
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication 
from os.path import basename, isfile
from bs4 import BeautifulSoup as bs
from openpyxl import Workbook, load_workbook, utils


# make request
def get_data(currency, moment_start, moment_end):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:109.0) \
Gecko/20100101 Firefox/109.0"
    }

    url = f"https://www.moex.com/export/derivatives/currency-rate.aspx?\
language=ru&currency={currency}&moment_start={moment_start}&moment_end={moment_end}"

    response = requests.get(url, headers = headers)

    return response.text


# parse response into dictionary by columns
def parse_response_to_dict(response_text):
    soup = bs(response_text, features="lxml-xml")    
    rates = soup.find_all("rate")
    currency_dict = {"date":[], "value":[], "time":[]}

    for rate in rates:
        date, time = rate["moment"].split(" ")
        # add only evening clearing
        if int(time[:2]) < 18:
            continue
        year, month, day = date.split("-")
        hour, minute, second = time.split(":")
        currency_dict["date"].append(datetime.date(int(year),int(month),int(day)))
        currency_dict["value"].append(float(rate["value"]))
        currency_dict["time"].append(datetime.time(int(hour),int(minute),int(second))
        )

    return currency_dict


# make Excel table
def make_xlsx(data_dict, report_name):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # note: As of Python version 3.11, dictionaries are ordered
    # so iterating through dict will be consistent 
    iterator = iter(data_dict)
    currency1 = next(iterator)
    currency2 = next(iterator)
    # column names
    ws.append([
        f"Дата {currency1}",
        f"Курс {currency1}",
        f"Время {currency1}",
        f"Дата {currency2}",
        f"Курс {currency2}",
        f"Время {currency2}",
        "Результат"
        ])

    # add parsed data
    col = 1
    for currency in data_dict:
       for key in data_dict[currency]:
            row = 2
            for elem in data_dict[currency][key]:
                ws.cell(column=col,row=row).value = elem
                row += 1
            col += 1

    # add results column
    excel_row_num = row - 1
    for row in range(2, excel_row_num+1):
        ws[f"G{row}"].value = f"=B{row}/E{row}"

    # format document
    column_letters = tuple(
        utils.get_column_letter(col_number + 1)
        for col_number in range(ws.max_column + 1)
        )

    for column_letter in column_letters:
        ws.column_dimensions[column_letter].bestFit = True

    wb.save(report_name)


# make text for message body
def gen_email_body_text(num):
    text = f"Сгенерирован отчет размером {num} "
    if (num%10) == 1 and num != 11:
        text += "строка."
    elif (num%10) in [2,3,4] and num not in [12,13,14]:
        text += "строки."
    else:
        text += "строк."
    return text


# send email with attachment
def send_email(params_path, attachment_path):    
    params = {}
    row_number = 0
    with open(params_path, "r") as params_file:
        params = json.loads(params_file.read())

    msg = MIMEMultipart()
    msg["From"] = params["sender"]
    msg["To"] = params["recipient"]
    msg["Subject"] = "Ежемесячный отчет по индикативным курсам валют"

    with open(attachment_path, "rb") as file:
        attachment = MIMEApplication(
            file.read(), Name=basename(attachment_path))
        attachment["Content-Disposition"] = \
            f'attachment; filename="{basename(attachment_path)}"'

        wb = load_workbook(attachment_path)
        ws = wb.active
        row_number = ws.max_row

        msg.attach(MIMEText(gen_email_body_text(row_number),"plain"))
        msg.attach(attachment)

    context = ssl.create_default_context()
    with smtplib.SMTP_SSL(params["server"], params["port"], context=context) as server:
        server.login(params["sender"], params["password"])
        server.send_message(msg, from_addr=params["sender"], to_addrs=params["recipient"])

def main():
    # initialize variables
    currency1 = "USD/RUB"
    currency2 = "JPY/RUB"
    today = datetime.date.today()

    month_ago_end = today.replace(day=1) - datetime.timedelta(days=1)
    month_ago_start = month_ago_end.replace(day=1)

    moment_start = f"{month_ago_start.year}-{month_ago_start.month}-{month_ago_start.day}"
    moment_end = f"{month_ago_end.year}-{month_ago_end.month}-{month_ago_end.day}"
    moment = f"{month_ago_start.year}-{month_ago_start.month}"

    report_name = f"Report_{moment}.xlsx"
    attachment_path = f"./{report_name}"
    params_path = "./mailing_params.json"

    data_dict = dict.fromkeys([currency1, currency2])
    
    try:
        print(f"Checking for {report_name}")

        # check if report exists
        if not isfile(attachment_path):
            print(f"Not found\nRequesting data from server")
            # get data from server
            for currency in data_dict:
                data = get_data(currency, moment_start, moment_end)
                data_dict[currency] = parse_response_to_dict(data)
            print("Server request successfull")

            # make xlsx file
            make_xlsx(data_dict, report_name)
            print("Report generation successfull")
            
        else:
            print("Found!")   
        # send email
        # print(f"Sending {report_name}...")
        # send_email(params_path, attachment_path)
        # print("Success")

    except:
        print(f"########\n{traceback.format_exc()}########\n")

if __name__ == "__main__":
    main()